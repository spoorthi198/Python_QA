import openpyxl as xl

wb = xl.load_workbook('excel.xlsx')
sheet = wb['Sheet2']
cell = sheet.cell(1,1)
print(cell.value)
for row in range(1,sheet.max_row+1):
    for col in range(1,sheet.max_column+1):
        c = sheet.cell(row,col)
        print(c.value)

for r in sheet["A1":"B10"]:
     for c in r:
         print(c.value)
